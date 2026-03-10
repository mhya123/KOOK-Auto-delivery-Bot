from __future__ import annotations

from collections.abc import Iterable
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook


def build_recharge_cards_workbook(rows: Iterable[dict[str, object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "recharge_cards"
    sheet.append(
        [
            "code",
            "amount",
            "is_used",
            "used_by",
            "used_at",
            "created_by",
            "created_at",
        ]
    )
    for row in rows:
        sheet.append(
            [
                row.get("code", ""),
                int(row.get("amount", 0) or 0),
                "yes" if int(row.get("is_used", 0) or 0) == 1 else "no",
                row.get("used_by", "") or "",
                _format_timestamp(row.get("used_at")),
                row.get("created_by", "") or "",
                _format_timestamp(row.get("created_at")),
            ]
        )
    return _workbook_to_bytes(workbook)


def build_product_keys_workbook(grouped_rows: dict[str, list[dict[str, object]]]) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Group product key exports by worksheet so all-product exports stay readable.
    used_sheet_names: set[str] = set()
    for group_name, rows in grouped_rows.items():
        sheet = workbook.create_sheet(title=_unique_sheet_name(group_name, used_sheet_names))
        sheet.append(
            [
                "product_id",
                "product_name",
                "key_id",
                "key_content",
                "price",
                "is_sold",
                "sold_to",
                "sold_at",
                "created_by",
                "created_at",
            ]
        )
        for row in rows:
            sheet.append(
                [
                    int(row.get("product_id", 0) or 0),
                    row.get("product_name", "") or "",
                    int(row.get("key_id", 0) or 0),
                    row.get("key_content", "") or "",
                    int(row.get("price", 0) or 0),
                    "yes" if int(row.get("is_sold", 0) or 0) == 1 else "no",
                    row.get("sold_to", "") or "",
                    _format_timestamp(row.get("sold_at")),
                    row.get("created_by", "") or "",
                    _format_timestamp(row.get("created_at")),
                ]
            )
    return _workbook_to_bytes(workbook)


def _workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _format_timestamp(value: object) -> str:
    if value in {None, "", 0, "0"}:
        return ""
    try:
        timestamp = int(value)
    except (TypeError, ValueError):
        return str(value)
    return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")


def _unique_sheet_name(raw_name: str, used_sheet_names: set[str]) -> str:
    base_name = _sanitize_sheet_name(raw_name) or "sheet"
    if base_name not in used_sheet_names:
        used_sheet_names.add(base_name)
        return base_name

    index = 2
    while True:
        suffix = f"_{index}"
        candidate = f"{base_name[: 31 - len(suffix)]}{suffix}"
        if candidate not in used_sheet_names:
            used_sheet_names.add(candidate)
            return candidate
        index += 1


def _sanitize_sheet_name(raw_name: str) -> str:
    invalid_chars = set('[]:*?/\\')
    cleaned = "".join("_" if char in invalid_chars else char for char in raw_name).strip()
    return cleaned[:31]
